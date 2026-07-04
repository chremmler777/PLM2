"""API endpoints for revision-scoped part BOMs - Phase 5.

A sub-assembly's revision owns its bill of materials, so frozen revisions
keep the exact component list they were approved with. Items reference a
project part, a catalog part, or are free text.
"""
import logging
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Depends, status
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import joinedload

from app.dependencies import get_current_user
from app.models import get_db
from app.models import User, PartBOMItem
from app.models.catalog import CatalogPart
from app.models.part import Part, RevisionStatus
from app.services.part_service import PartService, RevisionService, ChangelogService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/parts", tags=["part-bom"])

LOCKED_STATUSES = {
    RevisionStatus.FROZEN.value,
    RevisionStatus.CANCELLED.value,
    RevisionStatus.ARCHIVED.value,
}


def _status_value(rev_status) -> str:
    return rev_status.value if hasattr(rev_status, "value") else str(rev_status)


class BOMItemCreate(BaseModel):
    child_part_id: Optional[int] = None
    catalog_part_id: Optional[int] = None
    name: Optional[str] = Field(None, max_length=255, description="Required for free-text items")
    quantity: float = Field(1.0, gt=0)
    unit: str = Field("pcs", max_length=20)
    notes: Optional[str] = None


class BOMItemUpdate(BaseModel):
    quantity: Optional[float] = Field(None, gt=0)
    unit: Optional[str] = Field(None, max_length=20)
    position: Optional[int] = None
    notes: Optional[str] = None


def _item_dict(item: PartBOMItem, child_part: Optional[Part] = None, catalog_part: Optional[CatalogPart] = None) -> dict:
    return {
        "id": item.id,
        "revision_id": item.revision_id,
        "child_part_id": item.child_part_id,
        "catalog_part_id": item.catalog_part_id,
        "item_number": item.item_number,
        "name": item.name,
        "quantity": item.quantity,
        "unit": item.unit,
        "position": item.position,
        "notes": item.notes,
        "child_part_number": child_part.part_number if child_part else None,
        "child_part_type": child_part.part_type if child_part else None,
        "catalog_part_number": catalog_part.part_number if catalog_part else None,
        "catalog_supplier": catalog_part.supplier if catalog_part else None,
    }


async def _get_unlocked_revision(db: AsyncSession, revision_id: int):
    revision = await RevisionService.get_revision(db, revision_id)
    if not revision:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if _status_value(revision.status) in LOCKED_STATUSES:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Revision {revision.revision_name} is {_status_value(revision.status)}; its BOM is read-only",
        )
    return revision


@router.get("/revisions/{revision_id}/bom", response_model=List[dict])
async def list_bom_items(
    revision_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List BOM items for a revision, with referenced part info resolved."""
    revision = await RevisionService.get_revision(db, revision_id)
    if not revision:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")

    result = await db.execute(
        select(PartBOMItem)
        .where(PartBOMItem.revision_id == revision_id)
        .options(joinedload(PartBOMItem.child_part))
        .order_by(PartBOMItem.position, PartBOMItem.id)
    )
    items = result.unique().scalars().all()

    catalog_ids = [i.catalog_part_id for i in items if i.catalog_part_id]
    catalog_map = {}
    if catalog_ids:
        cat_result = await db.execute(select(CatalogPart).where(CatalogPart.id.in_(catalog_ids)))
        catalog_map = {c.id: c for c in cat_result.scalars().all()}

    return [_item_dict(i, i.child_part, catalog_map.get(i.catalog_part_id)) for i in items]


@router.get("/revisions/{revision_id}/bom/export")
async def export_bom_xlsx(
    revision_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Export a revision's BOM as an Excel workbook."""
    import io
    from datetime import datetime

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    revision = await RevisionService.get_revision(db, revision_id)
    if not revision:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    part = await PartService.get_part(db, revision.part_id)

    result = await db.execute(
        select(PartBOMItem)
        .where(PartBOMItem.revision_id == revision_id)
        .options(joinedload(PartBOMItem.child_part))
        .order_by(PartBOMItem.position, PartBOMItem.id)
    )
    items = result.unique().scalars().all()

    catalog_ids = [i.catalog_part_id for i in items if i.catalog_part_id]
    catalog_map = {}
    if catalog_ids:
        cat_result = await db.execute(select(CatalogPart).where(CatalogPart.id.in_(catalog_ids)))
        catalog_map = {c.id: c for c in cat_result.scalars().all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    ws["A1"] = f"Bill of Materials — {part.name} ({part.part_number})"
    ws["A1"].font = title_font
    ws["A2"] = f"Revision: {revision.revision_name}"
    ws["A3"] = f"Exported: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"

    headers = ["Pos", "Item", "Part Number", "Qty", "Unit", "Supplier", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, item in enumerate(items, start=6):
        catalog = catalog_map.get(item.catalog_part_id)
        ws.cell(row=row, column=1, value=item.item_number)
        ws.cell(row=row, column=2, value=item.name)
        ws.cell(row=row, column=3, value=(
            item.child_part.part_number if item.child_part
            else catalog.part_number if catalog else ""
        ))
        ws.cell(row=row, column=4, value=item.quantity)
        ws.cell(row=row, column=5, value=item.unit)
        ws.cell(row=row, column=6, value=catalog.supplier if catalog and catalog.supplier else "")
        ws.cell(row=row, column=7, value=item.notes or "")

    widths = [8, 32, 18, 8, 8, 20, 36]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=5, column=col).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"BOM_{part.part_number}_{revision.revision_name}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{part_id}/revisions/{revision_id}/bom", response_model=dict, status_code=status.HTTP_201_CREATED)
async def add_bom_item(
    part_id: int,
    revision_id: int,
    body: BOMItemCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Add a BOM item to a sub-assembly's revision."""
    try:
        part = await PartService.get_part(db, part_id)
        if not part:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Part not found")

        revision = await _get_unlocked_revision(db, revision_id)
        if revision.part_id != part_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Revision does not belong to this part",
            )

        # Resolve the item's display name from its reference
        child_part = None
        catalog_part = None
        if body.child_part_id:
            if body.child_part_id == part_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="A part cannot be a BOM item of itself",
                )
            child_part = await PartService.get_part(db, body.child_part_id)
            if not child_part:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Child part not found")
            name = child_part.name
        elif body.catalog_part_id:
            cat_result = await db.execute(select(CatalogPart).where(CatalogPart.id == body.catalog_part_id))
            catalog_part = cat_result.scalar_one_or_none()
            if not catalog_part:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catalog part not found")
            name = catalog_part.name
        elif body.name:
            name = body.name
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Provide child_part_id, catalog_part_id, or a name",
            )

        # Next position / item number
        max_pos = (
            await db.execute(
                select(func.max(PartBOMItem.position)).where(PartBOMItem.revision_id == revision_id)
            )
        ).scalar() or 0
        next_pos = max_pos + 10

        item = PartBOMItem(
            revision_id=revision_id,
            child_part_id=body.child_part_id,
            catalog_part_id=body.catalog_part_id,
            item_number=str(next_pos),
            name=name,
            quantity=body.quantity,
            unit=body.unit,
            position=next_pos,
            notes=body.notes,
            created_by=current_user.id,
        )
        db.add(item)
        await db.flush()

        await ChangelogService.log_action(
            db,
            part_id=part_id,
            revision_id=revision_id,
            action="bom_item_added",
            action_description=f"Added BOM item '{name}' (qty {body.quantity} {body.unit}) to {revision.revision_name}",
            performed_by=current_user.id,
        )
        await db.commit()

        return _item_dict(item, child_part, catalog_part)
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Failed to add BOM item to revision {revision_id}: {e}", exc_info=True)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@router.put("/bom-items/{item_id}", response_model=dict)
async def update_bom_item(
    item_id: int,
    body: BOMItemUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update quantity/unit/position/notes of a BOM item."""
    try:
        result = await db.execute(
            select(PartBOMItem).where(PartBOMItem.id == item_id).options(joinedload(PartBOMItem.child_part))
        )
        item = result.unique().scalar_one_or_none()
        if not item:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BOM item not found")

        revision = await _get_unlocked_revision(db, item.revision_id)

        changes = []
        for field in ("quantity", "unit", "position", "notes"):
            value = getattr(body, field)
            if value is not None and value != getattr(item, field):
                changes.append(f"{field}: {getattr(item, field)} -> {value}")
                setattr(item, field, value)

        if changes:
            await ChangelogService.log_action(
                db,
                part_id=revision.part_id,
                revision_id=revision.id,
                action="bom_item_updated",
                action_description=f"Updated BOM item '{item.name}' ({'; '.join(changes)})",
                performed_by=current_user.id,
            )
        await db.commit()
        return _item_dict(item, item.child_part)
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Failed to update BOM item {item_id}: {e}", exc_info=True)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@router.delete("/bom-items/{item_id}")
async def delete_bom_item(
    item_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Remove a BOM item from its revision."""
    try:
        result = await db.execute(select(PartBOMItem).where(PartBOMItem.id == item_id))
        item = result.scalar_one_or_none()
        if not item:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BOM item not found")

        revision = await _get_unlocked_revision(db, item.revision_id)

        await ChangelogService.log_action(
            db,
            part_id=revision.part_id,
            revision_id=revision.id,
            action="bom_item_removed",
            action_description=f"Removed BOM item '{item.name}' from {revision.revision_name}",
            performed_by=current_user.id,
        )
        await db.delete(item)
        await db.commit()
        return {"status": "success", "message": "BOM item removed"}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Failed to delete BOM item {item_id}: {e}", exc_info=True)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))
