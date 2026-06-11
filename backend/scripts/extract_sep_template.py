"""One-off extraction: GB-DP-0001 SEP matrix xlsm -> app/data/sep_template.json.

Reads the 7 gate sheets and produces the seed template used by "Activate SEP".
Re-run if the controlled document changes (new version of the matrix).
"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

XLSM = Path(__file__).resolve().parents[2] / "Documents" / "GB-DP-0001_SEP-Matrix_DE-EN_V01.xlsm"
OUT = Path(__file__).resolve().parents[1] / "app" / "data" / "sep_template.json"

GATE_SHEETS = [
    ("K0/RG1", "K0_RG1"),
    ("K/RG2", "K_RG2"),
    ("E/RG3", "E_RG3"),
    ("D/RG4", "D_RG4"),
    ("C/RG5", "C_RG5"),
    ("B/RG6", "B_RG6"),
    ("A/RG7", "A_RG7"),
]


def split_de_en(text: str) -> tuple[str, str]:
    """Titles look like 'German text \n(English text)'. Fall back to same text."""
    text = (text or "").strip()
    m = re.match(r"^(.*?)\s*[\n]\s*\((.*)\)\s*$", text, re.DOTALL)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return text, text


def main() -> None:
    wb = load_workbook(XLSM, read_only=True, data_only=True)
    gates = []
    for seq, (code, sheet_name) in enumerate(GATE_SHEETS, start=1):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=10, values_only=False))
        # Row 10 holds the gate phase description in column G ('RFQ - Nomination\n/ RFQ Nomination')
        phase_raw = None
        for row in rows[:3]:
            for cell in row:
                if cell.column_letter == "G" and cell.value:
                    phase_raw = str(cell.value)
                    break
            if phase_raw:
                break
        phase_de, phase_en = "", ""
        if phase_raw:
            parts = [p.strip() for p in phase_raw.replace("\n", " ").split("/") if p.strip()]
            if len(parts) >= 2:
                phase_de, phase_en = parts[0], parts[-1]
            else:
                phase_de = phase_en = phase_raw.strip()

        items = []
        for row in ws.iter_rows(min_row=14):
            no = row[1].value  # B: item number
            title = row[2].value  # C: task DE/(EN)
            psp = row[4].value  # E: PSP no
            dept = row[5].value  # F: responsible department
            if no is None or not str(no).strip().isdigit():
                continue
            title_de, title_en = split_de_en(str(title) if title else "")
            if not title_de:
                continue
            items.append({
                "item_no": int(str(no).strip()),
                "title_de": title_de,
                "title_en": title_en,
                "psp_no": str(psp).strip() if psp is not None else None,
                "department": str(dept).strip() if dept else "(to be defined)",
            })
        gates.append({
            "code": code,
            "seq": seq,
            "phase_de": phase_de,
            "phase_en": phase_en,
            "items": items,
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    total = sum(len(g["items"]) for g in gates)
    OUT.write_text(json.dumps({"source": XLSM.name, "gates": gates}, ensure_ascii=False, indent=1))
    print(f"Wrote {OUT}: {len(gates)} gates, {total} work items")
    for g in gates:
        print(f"  {g['code']}: {len(g['items'])} items | {g['phase_en']}")


if __name__ == "__main__":
    main()
